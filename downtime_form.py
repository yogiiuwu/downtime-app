
import streamlit as st
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from copy import copy
from PIL import Image
import tempfile
import re
import gspread
import json
import ast
from oauth2client.service_account import ServiceAccountCredentials
import hashlib
import pandas as pd
import shutil
import os

# --- Load login state dari file jika session kosong ---
if "logged_in" not in st.session_state:
    try:
        with open("login_state.json", "r") as f:
            data = json.load(f)
            st.session_state.logged_in = data.get("logged_in", False)
            st.session_state.username = data.get("username", "")
    except FileNotFoundError:
        st.session_state.logged_in = False
        st.session_state.username = ""

# LOGIN SECTION
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

users = {
    "admin": hash_password("admin"),
    "yogi": hash_password("8081"),
    "arfian": hash_password("2178"),
    "cakrahayu": hash_password("cakrahayu2003"),
    "daffa": hash_password("8058"),
    "rokhim": hash_password("2090"),
    "sheva": hash_password("2175")
}

def check_login(username, password):
    return username in users and users[username] == hash_password(password)

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
# Tambahkan halaman reset password
def reset_password():
    st.title("🔑 Reset Password")

    username = st.text_input("Masukkan Username")

    if st.button("Cek Username"):
        if username in users:
            st.session_state.reset_user = username
            st.session_state.step_reset = "new_password"
            st.rerun()
        else:
            st.error("Username tidak ditemukan!")

# Form input password baru
def input_password_baru():
    st.title("🔑 Buat Password Baru")

    new_password = st.text_input("Masukkan Password Baru", type="password")
    confirm_password = st.text_input("Ulangi Password Baru", type="password")

    if st.button("Simpan Password Baru"):
        if new_password != confirm_password:
            st.error("Password tidak sama, ulangi lagi.")
        elif new_password == "":
            st.error("Password tidak boleh kosong.")
        else:
            # Simpan password baru (di-hash)
            users[st.session_state.reset_user] = hash_password(new_password)
            st.success("Password berhasil diubah, silakan login.")
            st.session_state.reset_user = None
            st.session_state.step_reset = None
            st.rerun()

if "step_reset" not in st.session_state:
    st.session_state.step_reset = None

if st.session_state.step_reset == "reset":
    reset_password()
    st.stop()

if st.session_state.step_reset == "new_password":
    input_password_baru()
    st.stop()

if not st.session_state.logged_in:
    st.title("🔐 Login")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    col_login, col_lupa = st.columns(2)
    with col_login:
        if st.button("Login"):
            if check_login(username, password):
                st.session_state.logged_in = True
                st.session_state.username = username

                # Simpan ke file login_state.json
                with open("login_state.json", "w") as f:
                    json.dump({
                        "logged_in": True,
                        "username": username
                    }, f)

                st.success("Login berhasil!")
                st.rerun()

            else:
                st.error("Username atau password salah")

    with col_lupa:
        if st.button("Lupa Password?"):
            st.session_state.step_reset = "reset"
            st.rerun()

    st.stop()

# GOOGLE SHEETS
def get_google_sheet(sheet_name):
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds_dict = json.loads(json.dumps(dict(st.secrets["gcp_service_account"])))
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    client = gspread.authorize(creds)
    spreadsheet = client.open(sheet_name)
    return spreadsheet

def check_login(username, password):
    return username in users and users[username] == hash_password(password)

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "show_summary" not in st.session_state:
    st.session_state.show_summary = False

if not st.session_state.logged_in:
    st.title("🔐 Login")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login"):
        if check_login(username, password):
            st.session_state.logged_in = True
            st.success("Login berhasil!")
            st.rerun()
        else:
            st.error("Username atau password salah")
    st.stop()


def normalize(text):
    return re.sub(r"\s+", " ", str(text).strip()).lower()
def find_or_create_lot_block(ws, lot, template_start_row=8, template_height=18):
    max_row = ws.max_row
    lot_column = "C"

    # === CARI APAKAH LOT SUDAH ADA ===
    for row in range(template_start_row, max_row + 1, template_height + 2):
        lot_cell = ws[f"{lot_column}{row + 2}"]
        if lot_cell.value and str(lot_cell.value).strip().lower() == lot.lower():
            return row

    # === CARI BLOK KOSONG ===
    for row in range(template_start_row, max_row + 1, template_height + 2):
        if not (ws.cell(row=row, column=3).value or
                ws.cell(row=row + 1, column=3).value or
                ws.cell(row=row + 2, column=3).value):
            new_start = row
            break
    else:
        # === SALIN TEMPLATE BARU ===
        new_start = max_row + 2
        for i in range(template_height):
            src_row = template_start_row + i
            dst_row = new_start + i
            for col in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=src_row, column=col)
                dst_cell = ws.cell(row=dst_row, column=col)
                if not isinstance(src_cell, MergedCell):
                    dst_cell.value = src_cell.value
                if src_cell.has_style:
                    dst_cell._style = copy(src_cell._style)
                if src_cell.fill != PatternFill():
                    dst_cell.fill = copy(src_cell.fill)
                if src_cell.comment:
                    dst_cell.comment = Comment(src_cell.comment.text, "username")

        # SALIN MERGE TEMPLATE
        offset = new_start - template_start_row
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= template_start_row and rng.max_row <= template_start_row + template_height:
                ws.merge_cells(start_row=rng.min_row + offset,
                               end_row=rng.max_row + offset,
                               start_column=rng.min_col,
                               end_column=rng.max_col)

    # === NOMOR KOLOM A UNTUK DOWNTIME ===
    for i in range(11):
        cell = ws.cell(row=new_start + 5 + i, column=1)
        if not isinstance(cell, MergedCell):
            cell.value = i + 1

    # === FORMULA TOTAL ===
    ws[f"AC{new_start + 16}"] = f"=SUM(AC{new_start + 5}:AC{new_start + 15})"
    ws[f"AD{new_start + 16}"] = f"=SUM(AD{new_start + 5}:AD{new_start + 15})"

    # === HITUNG BLOK DENGAN RANGE MERGE DI KOLOM A ===
    blok_nomor = 1
    for rng in ws.merged_cells.ranges:
        if rng.min_col == 1 and rng.max_col == 1:
            if (rng.max_row - rng.min_row + 1) == template_height:
                if rng.max_row < new_start:
                    blok_nomor += 1

    # === MERGE & ISI NOMOR BLOK ===
    ws.merge_cells(start_row=new_start, end_row=new_start + template_height - 1,
                   start_column=1, end_column=1)
    ws.cell(row=new_start, column=1).value = blok_nomor

    return new_start
def isi_metadata_ke_semua_sheet(file_path, metadata, username):
    wb = load_workbook(file_path)

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        try:
            def isi(cell_pos, value, comment_title):
                cell = ws[cell_pos]
                if not isinstance(cell, MergedCell):
                    cell.value = value

            isi("C8", metadata["nama_produk"], "Nama Produk")
            isi("C9", metadata["kode_produk"], "Kode Produk")
            isi("C10", metadata["lot"], "Kode LOT")
            isi("C11", str(metadata["tanggal_produksi"]), "Tanggal Produksi")

        except Exception as e:
            print(f"⚠️ Error isi metadata sheet {sheet}: {e}")

    wb.save(file_path)

# Simpan ke Excel 
def simpan_downtime_ke_excel(template_path, metadata, entry):
    wb = load_workbook(template_path)

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        def set_meta_if_not_merged(cell_key, value, comment_text):
            cell = ws[cell_key]
            if not isinstance(cell, MergedCell):
                cell.value = value
                cell.comment = Comment(comment_text, st.session_state.username)

        set_meta_if_not_merged("C1", metadata["nama_produk"], "Nama Produk")
        set_meta_if_not_merged("C2", metadata["kode_produk"], "Kode Produk")
        set_meta_if_not_merged("C3", metadata["lot"], "Kode LOT")
        set_meta_if_not_merged("C4", str(metadata["tanggal_produksi"]), "Tanggal Produksi")

    # Sheet utama tempat menulis downtime
    sheet_name = metadata["line_produksi"]
    ws = wb[sheet_name]

    blok_awal = find_or_create_lot_block(ws, metadata["lot"])

    # Tulis metadata lagi di atas blok
    ws[f"C{blok_awal}"] = metadata["nama_produk"]
    ws[f"C{blok_awal + 1}"] = metadata["kode_produk"]
    ws[f"C{blok_awal + 2}"] = metadata["lot"]
    ws[f"C{blok_awal + 3}"] = str(metadata["tanggal_produksi"])

    # Tambahkan downtime
    jenis_downtime = normalize(entry["jenis"])
    jam_index = int(str(entry["jam"]).split(":")[0]) + 5
    durasi = float(entry["durasi"])
    durasi_sisa = durasi
    komentar = entry.get("komentar", "")
    found = False

    for i in range(5, 16):
        row = blok_awal + i
        cell_value = ws.cell(row=row, column=4).value
        if normalize(cell_value) == jenis_downtime:
            while durasi_sisa > 0:
                durasi_input = min(durasi_sisa, 60)
                jam_col = jam_index
                cell = ws.cell(row=row, column=jam_col)
                if cell.value:
                    cell.value += durasi_input
                else:
                    cell.value = durasi_input
                if komentar:
                    cell.comment = Comment(komentar, st.session_state["username"])
                durasi_sisa -= durasi_input
                jam_index += 1  # ke jam berikutnya
            total_menit = sum(float(ws.cell(row=row, column=col).value or 0) for col in range(5, 29))
            ws.cell(row=row, column=29).value = total_menit
            ws.cell(row=row, column=30).value = round(total_menit / 60, 2)
            found = True
            break

    if not found:
        st.error(f"❌ Jenis downtime '{entry['jenis']}' tidak ditemukan pada LOT ini.")
       
    wb.save(template_path)

# ======================= STREAMLIT APP ========================

st.set_page_config(page_title="DOWNTIME SOFTBAG II", layout="wide")
# Tambahkan tombol logout
with st.sidebar:
    st.image("otsuka_logo.png", width=100)

    if not st.session_state.logged_in:
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        if st.button("🔓 Login"):         
            if username and password:
                st.session_state.logged_in = True
                st.session_state.username = username
                st.experimental_rerun()
    else:
        st.markdown(f"👤 **{st.session_state.username}**")
        if st.button("🔒 Logout"):
            st.session_state.logged_in = False
            st.session_state.username = ""

            # Hapus file login_state.json saat logout
            if os.path.exists("login_state.json"):
                os.remove("login_state.json")

            st.rerun()

st.title(f"Form Input Downtime Packing (User: {st.session_state.username})")

if "excel_path" not in st.session_state:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    with open("template_downtime_multi.xlsx", "rb") as fsrc:
        tmp.write(fsrc.read())
    tmp.close()
    st.session_state.excel_path = tmp.name

if "updated_excel" not in st.session_state:
    with open(st.session_state.excel_path, "rb") as f:
        st.session_state.updated_excel = f.read()

if "history_downtime" not in st.session_state:
    st.session_state.history_downtime = []

logo = Image.open("otsuka_logo.png")
col1, col2 = st.columns([1, 10])
with col1:
    st.image(logo, width=60)
with col2:
    st.markdown("## Form Input Downtime Packing")

line_options = [
    "DT ALT A", "DT ALT B", "DT Autocase A", "DT Autocase B",
    "DT Carton Erector", "DT Carton Sealing A", "DT Carton Sealing B"
]
downtime_mapping = {
    "DT ALT A": {
        "Utility Downtime": ["Pressure Air Drop", "Listrik Padam"],
        "Proses Downtime": ["Setting Mesin", "Posisi Produk Abnormal"],
        "Mesin ALT Downtime": [
            "Parts Electrical / Control Error", "Parts Mechanical Error",
            "Parts Pneumatic Error", "Robot Spider", "No Operator"
        ]
    },
    "DT ALT B": {
        "Utility Downtime": ["Pressure Air Drop", "Listrik Padam"],
        "Proses Downtime": ["Setting Mesin", "Posisi Produk Abnormal"],
        "Mesin ALT Downtime": [
            "Parts Electrical / Control Error", "Parts Mechanical Error",
            "Parts Pneumatic Error", "Robot Spider", "No Operator"
        ]
    },
    "DT Autocase A": {
        "Utility Downtime": ["Pressure Air Drop", "Listrik Padam"],
        "Proses Downtime": ["Inspeksi Proses", "Material Habis", "Ganti Ribbon/Label"],
        "Mesin ALT Downtime": [
            "Parts Electrical / Control Error", "Parts Mechanical Error",
            "Parts Pneumatic Error", "Conveyor Transfer Trouble", "Motor Trouble", "No Operator"
        ]
    },
    "DT Autocase B": {
        "Utility Downtime": ["Pressure Air Drop", "Listrik Padam"],
        "Proses Downtime": ["Inspeksi Proses", "Material Habis", "Ganti Ribbon/Label"],
        "Mesin ALT Downtime": [
            "Parts Electrical / Control Error", "Parts Mechanical Error",
            "Parts Pneumatic Error", "Conveyor Transfer Trouble", "Motor Trouble", "No Operator"
        ]
    },
    "DT Carton Erector": {
        "Utility Downtime": ["Pressure Air Drop", "Listrik Padam", "Nitrogen Supply"],
        "Proses Downtime": ["Menunggu Produk Sterilisasi", "Inspeksi Proses", "Material Habis", "Penggantian Material"],
        "Mesin Weight Checker Downtime": [
            "Parts Electrical / Control Error", "Parts Mechanical Error",
            "Parts Pneumatic Error", "Conveyor Error", "Motor Error", "No Operator"
        ]
    },
    "DT Carton Sealing A": {
        "Utility Downtime": ["Pressure Air Drop", "Listrik Padam"],
        "Proses Downtime": ["Menunggu Produk Sterilisasi", "Inspeksi Proses", "Material Habis"],
        "Mesin Weight Checker Downtime": [
            "Parts Electrical / Control Error", "Parts Mechanical Error",
            "Parts Pneumatic Error", "Conveyor Error", "Motor Error", "No Operator"
        ]
    },
    "DT Carton Sealing B": {
        "Utility Downtime": ["Pressure Air Drop", "Listrik Padam"],
        "Proses Downtime": ["Menunggu Produk Sterilisasi", "Inspeksi Proses", "Material Habis"],
        "Mesin Weight Checker Downtime": [
            "Parts Electrical / Control Error", "Parts Mechanical Error",
            "Parts Pneumatic Error", "Conveyor Error", "Motor Error", "No Operator"
        ]
    }
}

# Fungsi untuk simpan ke Google Sheet
def simpan_downtime_ke_sheet(sheet, metadata, entry):
    sheet.append_row([
        str(datetime.now()),  
        # Timestamp
        metadata["line_produksi"],
        metadata["nama_produk"],
        metadata["kode_produk"],
        metadata["lot"],
        str(metadata["tanggal_produksi"]),
        entry["jenis"],
        entry["jam"],
        entry["durasi"],
        entry["komentar"]
    ])

# ====== FORM STREAMLIT ======
# Di luar st.form() agar bisa dinamis:
line_produksi = st.selectbox("Pilih Jenis Mesin", line_options, key="line_produksi")
opsi = downtime_mapping.get(line_produksi, {})
kategori = st.selectbox("Kategori Downtime", list(opsi.keys()), key="kategori")
jenis = st.selectbox("Jenis Downtime", opsi.get(kategori, []), key="jenis")

# Sekarang form mulai:
with st.form("form_downtime"):
    st.subheader("📦 Form Input Downtime")

    col1, col2, col3 = st.columns(3)
    with col1:
        nama_produk = st.text_input("Nama Produk")
    with col2:
        kode_produk = st.text_input("Kode Produk")
    with col3:
        lot = st.text_input("Kode LOT")

    tgl_produksi = st.date_input("Tanggal Produksi", value=datetime.today())

    col4, col5 = st.columns(2)
    with col4:
        jam = st.selectbox("Jam Terjadi", [f"{j:02d}:00" for j in range(24)])
    with col5:
        durasi = st.number_input("Durasi (menit)", min_value=1)

    komentar = st.text_area("Komentar", placeholder="tambahkan keterangan")

    submitted = st.form_submit_button("➕ Tambahkan Downtime")
if submitted:
    if not nama_produk or not kode_produk or not lot:
        st.warning("⚠️ Harap isi semua data produk.")
    else:
        metadata = {
            "nama_produk": nama_produk,
            "kode_produk": kode_produk,
            "lot": lot,
            "tanggal_produksi": tgl_produksi,
            "line_produksi": line_produksi
        }
        entry = {
            "jenis": jenis,
            "jam": jam,
            "durasi": durasi,
            "komentar": komentar
        }

        # 🟢 Tambahkan baris ini
        isi_metadata_ke_semua_sheet(st.session_state.excel_path, metadata, st.session_state.username)
        simpan_downtime_ke_excel(st.session_state.excel_path, metadata, entry)
        try:
            gsheet = get_google_sheet("DATABASE")
            simpan_downtime_ke_sheet(gsheet.sheet1, metadata, entry)
            st.success("✅ Data berhasil disimpan ke Google Sheet!")
        except Exception as e:
            st.warning(f"Gagal simpan ke Google Sheet: {e}")

        with open(st.session_state.excel_path, "rb") as f:
            st.session_state.updated_excel = f.read()

        st.session_state.history_downtime.append(
            f"✅ {metadata['nama_produk']} (LOT: {metadata['lot']}) - {entry['durasi']} menit ditambahkan.")

col_tombol2, col_tombol3 = st.columns([2, 1])
with col_tombol2:
    with open(st.session_state.excel_path, "rb") as f:
        excel_bytes = f.read()
    filename = f"DT_{nama_produk.replace(' ', '_').upper()}_{lot.upper()}.xlsx"
    if st.download_button("📥 Download Excel", data=excel_bytes, file_name=filename,
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
        st.session_state.history_downtime.clear()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        with open("template_downtime_multi.xlsx", "rb") as fsrc:
            tmp.write(fsrc.read())
        tmp.close()
        st.session_state.excel_path = tmp.name
with col_tombol3:
    if st.button("📊 Summary"):
        st.session_state.show_summary = True
if st.session_state.history_downtime:
    st.subheader("📋 Riwayat Downtime")
    for msg in st.session_state.history_downtime:
        st.success(msg)
        # ================= TAMBAHAN MENU DOWNLOAD =================
st.subheader("📥 Download Data Downtime Per Bulan")

try:
    gsheet = get_google_sheet("DATABASE")
    rows = gsheet.sheet1.get_all_records()
    df = pd.DataFrame(rows)
    
    if df.empty:
        st.warning("❌ Data Google Sheet masih kosong!")
    else:
        # Konversi tanggal produksi ke datetime
        df["Tanggal Produksi"] = pd.to_datetime(df["Tanggal Produksi"], errors='coerce')
        df = df.dropna(subset=["Tanggal Produksi"])  # Drop jika ada error parsing tanggal
        df["Bulan-Tahun"] = df["Tanggal Produksi"].dt.strftime('%Y-%m')

        bulan_options = sorted(df["Bulan-Tahun"].unique())
        selected_bulan = st.selectbox("Pilih Bulan Produksi:", bulan_options)
        if st.button("📥 Download Excel Bulanan"):
            filtered_df = df[df["Bulan-Tahun"] == selected_bulan]
            
            tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            shutil.copy("template_downtime_multi.xlsx", tmp_file.name)
            tmp_file.close()

            # Grouping per LOT
            lot_grouped = {}
            for _, row in filtered_df.iterrows():
                lot_key = row["Lot"]
                if lot_key not in lot_grouped:
                    lot_grouped[lot_key] = {
                        "metadata": {
                            "nama_produk": row["Nama Produk"],
                            "kode_produk": row["Kode Produk"],
                            "lot": row["Lot"],
                            "tanggal_produksi": row["Tanggal Produksi"].date(),
                            "line_produksi": row["Line Produksi"]
                        },
                        "entries": []
                    }
                lot_grouped[lot_key]["entries"].append({
                    "jenis": row["Jenis"],
                    "jam": row["Jam"],
                    "durasi": row["Durasi"],
                    "komentar": row["Komentar"]
                })

            for lot_data in lot_grouped.values():
                metadata = lot_data["metadata"]
                isi_metadata_ke_semua_sheet(tmp_file.name, metadata, st.session_state.username)
                for entry in lot_data["entries"]:
                    simpan_downtime_ke_excel(tmp_file.name, metadata, entry)

            # Download result
            with open(tmp_file.name, "rb") as f:
                excel_bytes = f.read()
            filename = f"Downtime_{selected_bulan}.xlsx"
            st.download_button("⬇️ Klik untuk Download", data=excel_bytes, file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            # Buat file excel baru dari template
            tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            shutil.copy("template_downtime_multi.xlsx", tmp_file.name)
            tmp_file.close()

            # Proses isi data per baris
            for idx, row in filtered_df.iterrows():
                metadata = {
                    "nama_produk": row["Nama Produk"],
                    "kode_produk": row["Kode Produk"],
                    "lot": row["Lot"],
                    "tanggal_produksi": row["Tanggal Produksi"].date(),
                    "line_produksi": row["Line Produksi"]
                }
                entry = {
                    "jenis": row["Jenis"],
                    "jam": row["Jam"],
                    "durasi": row["Durasi"],
                    "komentar": row["Komentar"]
                }
                simpan_downtime_ke_excel(tmp_file.name, metadata, entry)
            
            with open(tmp_file.name, "rb") as f:
                excel_bytes = f.read()
            filename = f"Downtime_{selected_bulan}.xlsx"
            st.download_button("⬇️ Klik untuk Download", data=excel_bytes, file_name=filename,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")                               
except Exception as e:
    st.error(f"❌ Gagal membaca Google Sheet: {e}")
    
# === SUMMARY POPUP ===
if st.session_state.show_summary:
    st.subheader("📊 Summary Downtime")

    try:
        gsheet = get_google_sheet("DATABASE")
        rows = gsheet.sheet1.get_all_records()
        df = pd.DataFrame(rows)
        
        if df.empty:
            st.warning("❌ Data Google Sheet masih kosong!")
            st.session_state.show_summary = False
        else:
            # Proses data
            df["Tanggal Produksi"] = pd.to_datetime(df["Tanggal Produksi"], errors='coerce')
            df = df.dropna(subset=["Tanggal Produksi"])
            df["Bulan-Tahun"] = df["Tanggal Produksi"].dt.strftime('%Y-%m')

            bulan_options = sorted(df["Bulan-Tahun"].unique())
            selected_bulan = st.selectbox("Pilih Bulan:", bulan_options)

            df_bulan = df[df["Bulan-Tahun"] == selected_bulan]
            summary = df_bulan.groupby("Line Produksi")["Durasi"].sum().reset_index()
            summary = summary.sort_values("Durasi", ascending=False)

            st.write("### Total Durasi per Line Produksi:")
            for idx, row in summary.iterrows():
                st.write(f"- **{row['Line Produksi']}**: {int(row['Durasi'])} menit")

            st.write("---")
            st.write("### Sample Komentar:")
            for mesin in df_bulan["Line Produksi"].unique():
                subdf = df_bulan[df_bulan["Line Produksi"] == mesin]
                komentar_sample = subdf["Komentar"].dropna().unique()[:3]
                st.write(f"**{mesin}**")
                for kom in komentar_sample:
                    st.write(f"- {kom}")

    except Exception as e:
        st.error(f"❌ Gagal membaca Google Sheet: {e}")

    if st.button("❌ Close Summary"):
        st.session_state.show_summary = False

if st.session_state.get("username") == "admin":
    st.subheader("🔄 Reset Downtime Per Bulan")

    try:
        gsheet = get_google_sheet("DATABASE")
        rows = gsheet.sheet1.get_all_records()
        df = pd.DataFrame(rows)

        if df.empty:
            st.warning("❌ Data Google Sheet kosong!")
        else:
            # Format tanggal
            df["Tanggal Produksi"] = pd.to_datetime(df["Tanggal Produksi"], errors='coerce')
            df = df.dropna(subset=["Tanggal Produksi"])
            df["Bulan-Tahun"] = df["Tanggal Produksi"].dt.strftime('%Y-%m')

            bulan_options = sorted(df["Bulan-Tahun"].unique())
            selected_bulan = st.selectbox("🗓 Pilih Bulan yang Akan Direset:", bulan_options)

            if st.button("🚨 Reset Downtime Bulan Ini"):
                st.session_state["confirm_reset"] = True

            if st.session_state.get("confirm_reset", False):
                with st.expander("⚠️ Konfirmasi Hapus Data Downtime", expanded=True):
                    st.warning(f"Apakah kamu yakin ingin menghapus seluruh data downtime bulan **{selected_bulan}**?")
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("✅ Ya, Reset Sekarang"):
                            try:
                                worksheet = gsheet.sheet1
                                all_data = worksheet.get_all_values()
                                headers = all_data[0]
                                rows_to_keep = [headers]

                                for row in all_data[1:]:
                                    tanggal = pd.to_datetime(row[5], errors="coerce")
                                    if pd.isna(tanggal) or tanggal.strftime('%Y-%m') != selected_bulan:
                                        rows_to_keep.append(row)

                                worksheet.clear()
                                worksheet.update("A1", rows_to_keep)
                                st.success(f"✅ Data bulan **{selected_bulan}** berhasil dihapus.")
                            except Exception as e:
                                st.error(f"❌ Gagal reset data: {e}")
                            st.session_state["confirm_reset"] = False

                    with col2:
                        if st.button("❌ Batal"):
                            st.session_state["confirm_reset"] = False

    except Exception as e:
        st.error(f"❌ Gagal membaca Google Sheet: {e}")
